from __future__ import annotations

from datetime import UTC, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .graph_service import clean_cell_value, create_graph_model


RELATION_HEADERS = [
    "Current(API)",
    "Current(Endpoint)",
    "Upstream(API)",
    "Upstream(Endpoint)",
    "Downstream(API)",
    "Downstream(Endpoint)",
    "Current(Business Group)",
    "Current(Business Group Source)",
    "Downstream(Business Group)",
    "Downstream(Business Group Source)",
]


def parse_workbook_bytes(data: bytes, source_name: str = "integration-metric.xlsx") -> dict[str, Any]:
    workbook = load_workbook(BytesIO(data), data_only=True)
    sheet_name = workbook.sheetnames[0] if workbook.sheetnames else None
    if not sheet_name:
        raise ValueError("Excel 文件里没有可读取的工作表。")
    sheet = workbook[sheet_name]
    rows: list[dict[str, Any]] = []
    for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if index <= 2:
            continue
        row_values = list(row or [])
        current_api = clean_cell_value(row_values[0] if len(row_values) > 0 else "")
        current_endpoint = clean_cell_value(row_values[1] if len(row_values) > 1 else "")
        upstream_api = clean_cell_value(row_values[2] if len(row_values) > 2 else "")
        upstream_endpoint = clean_cell_value(row_values[3] if len(row_values) > 3 else "")
        downstream_api = clean_cell_value(row_values[4] if len(row_values) > 4 else "")
        downstream_endpoint = clean_cell_value(row_values[5] if len(row_values) > 5 else "")
        if not current_api or not current_endpoint:
            continue
        rows.append({
            "id": len(rows) + 1,
            "currentApi": current_api,
            "currentEndpoint": current_endpoint,
            "upstreamApi": upstream_api,
            "upstreamEndpoint": upstream_endpoint,
            "downstreamApi": downstream_api,
            "downstreamEndpoint": downstream_endpoint,
            "currentBusinessGroup": clean_cell_value(row_values[6] if len(row_values) > 6 else ""),
            "currentBusinessGroupSource": clean_cell_value(row_values[7] if len(row_values) > 7 else ""),
            "downstreamBusinessGroup": clean_cell_value(row_values[8] if len(row_values) > 8 else ""),
            "downstreamBusinessGroupSource": clean_cell_value(row_values[9] if len(row_values) > 9 else ""),
            "upstreamBusinessGroup": "",
            "upstreamBusinessGroupSource": "",
        })
    return create_graph_model(rows, {"sourceName": source_name, "sheetName": sheet_name, "loadedAt": datetime.now(UTC).isoformat()})


def build_workbook_bytes(relations: list[dict[str, Any]], notes: dict[str, Any], title: str = "Generated from Mule repo scan") -> bytes:
    workbook = Workbook()
    relation_sheet = workbook.active
    if relation_sheet is None:
        relation_sheet = workbook.create_sheet("Relations", 0)
    assert isinstance(relation_sheet, Worksheet)
    relation_sheet.title = "Relations"
    relation_sheet.append([title])
    relation_sheet.append(RELATION_HEADERS)
    for relation in relations:
        relation_sheet.append([
            relation.get("currentApi", ""),
            relation.get("currentEndpoint", ""),
            relation.get("upstreamApi", ""),
            relation.get("upstreamEndpoint", ""),
            relation.get("downstreamApi", ""),
            relation.get("downstreamEndpoint", ""),
            relation.get("currentBusinessGroup", ""),
            relation.get("currentBusinessGroupSource", ""),
            relation.get("downstreamBusinessGroup", ""),
            relation.get("downstreamBusinessGroupSource", ""),
        ])
    notes_sheet = workbook.create_sheet("Notes")
    assert isinstance(notes_sheet, Worksheet)
    notes_sheet.append(["Repo", notes.get("repoName", "")])
    notes_sheet.append(["Resolved environment", notes.get("environment", "未解析")])
    notes_sheet.append(["Current API", notes.get("currentApi", "")])
    notes_sheet.append(["Business Group", notes.get("currentBusinessGroup", "未解析") or "未解析"])
    notes_sheet.append(["Relation count", str(len(relations))])
    notes_sheet.append(["Notes", notes.get("note", "Upstream caller systems are not inferred from a single repo unless another scanned repo calls this API.")])
    for index, summary in enumerate(notes.get("downstreamSummaries", []), start=1):
        notes_sheet.append([f"Downstream {index}", summary])
    widths = [26, 52, 22, 30, 30, 74, 28, 28, 28, 28]
    for index, width in enumerate(widths, start=1):
        relation_sheet.column_dimensions[chr(64 + index)].width = width
    notes_sheet.column_dimensions["A"].width = 24
    notes_sheet.column_dimensions["B"].width = 120
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
